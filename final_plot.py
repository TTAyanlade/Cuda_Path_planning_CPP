import numpy as np
import matplotlib.pyplot as plt
import math
import openpyxl
import time
import pandas as pd
from openpyxl import Workbook
import xlsxwriter
import statistics
# open excel file with collected data
data_rrt_blocked_11_expN_1= openpyxl.load_workbook('data_rrt_blocked_11_expN_1.xlsx')
data_rrt_blocked_11_expN_2= openpyxl.load_workbook('data_rrt_blocked_11_expN_2.xlsx')
data_rrt_blocked_11_expN_3= openpyxl.load_workbook('data_rrt_blocked_11_expN_3.xlsx')
data_rrt_blocked_11_expN_4= openpyxl.load_workbook('data_rrt_blocked_11_expN_4.xlsx')
data_rrt_blocked_11_expN_5= openpyxl.load_workbook('data_rrt_blocked_11_expN_5.xlsx')

data_rrt_blocked_11_expN_1_Sheet=data_rrt_blocked_11_expN_1.active
data_rrt_blocked_11_expN_2_Sheet=data_rrt_blocked_11_expN_2.active
data_rrt_blocked_11_expN_3_Sheet=data_rrt_blocked_11_expN_3.active
data_rrt_blocked_11_expN_4_Sheet=data_rrt_blocked_11_expN_4.active
data_rrt_blocked_11_expN_5_Sheet=data_rrt_blocked_11_expN_5.active

data_rrt_blocked_11_expN_1_Row=data_rrt_blocked_11_expN_1_Sheet.max_row
data_rrt_blocked_11_expN_2_Row=data_rrt_blocked_11_expN_2_Sheet.max_row
data_rrt_blocked_11_expN_3_Row=data_rrt_blocked_11_expN_3_Sheet.max_row
data_rrt_blocked_11_expN_4_Row=data_rrt_blocked_11_expN_4_Sheet.max_row
data_rrt_blocked_11_expN_5_Row=data_rrt_blocked_11_expN_5_Sheet.max_row

avrage_rrt_blocked_total_time=[]
avrage_rrt_blocked_cost=(data_rrt_blocked_11_expN_1_Sheet["C" + str(2)].value+data_rrt_blocked_11_expN_2_Sheet["C" + str(2)].value+data_rrt_blocked_11_expN_3_Sheet["C" + str(2)].value+data_rrt_blocked_11_expN_4_Sheet["C" + str(2)].value+data_rrt_blocked_11_expN_5_Sheet["C" + str(2)].value)/5
print('path length rrt_blocked')
print(avrage_rrt_blocked_cost)

data_rrt_gpu_blocked_11_expN_1= openpyxl.load_workbook('data_rrt_gpu_blocked_11_expN_1.xlsx')
data_rrt_gpu_blocked_11_expN_2= openpyxl.load_workbook('data_rrt_gpu_blocked_11_expN_2.xlsx')
data_rrt_gpu_blocked_11_expN_3= openpyxl.load_workbook('data_rrt_gpu_blocked_11_expN_3.xlsx')
data_rrt_gpu_blocked_11_expN_4= openpyxl.load_workbook('data_rrt_gpu_blocked_11_expN_4.xlsx')
data_rrt_gpu_blocked_11_expN_5= openpyxl.load_workbook('data_rrt_gpu_blocked_11_expN_5.xlsx')

data_rrt_gpu_blocked_11_expN_1_Sheet=data_rrt_gpu_blocked_11_expN_1.active
data_rrt_gpu_blocked_11_expN_2_Sheet=data_rrt_gpu_blocked_11_expN_2.active
data_rrt_gpu_blocked_11_expN_3_Sheet=data_rrt_gpu_blocked_11_expN_3.active
data_rrt_gpu_blocked_11_expN_4_Sheet=data_rrt_gpu_blocked_11_expN_4.active
data_rrt_gpu_blocked_11_expN_5_Sheet=data_rrt_gpu_blocked_11_expN_5.active

data_rrt_gpu_blocked_11_expN_1_Row=data_rrt_gpu_blocked_11_expN_1_Sheet.max_row
data_rrt_gpu_blocked_11_expN_2_Row=data_rrt_gpu_blocked_11_expN_2_Sheet.max_row
data_rrt_gpu_blocked_11_expN_3_Row=data_rrt_gpu_blocked_11_expN_3_Sheet.max_row
data_rrt_gpu_blocked_11_expN_4_Row=data_rrt_gpu_blocked_11_expN_4_Sheet.max_row
data_rrt_gpu_blocked_11_expN_5_Row=data_rrt_gpu_blocked_11_expN_5_Sheet.max_row

avrage_rrt_gpu_blocked_total_time=[]
avrage_rrt_gpu_blocked_cost=(data_rrt_gpu_blocked_11_expN_1_Sheet["C" + str(2)].value+data_rrt_gpu_blocked_11_expN_2_Sheet["C" + str(2)].value+data_rrt_gpu_blocked_11_expN_3_Sheet["C" + str(2)].value+data_rrt_gpu_blocked_11_expN_4_Sheet["C" + str(2)].value+data_rrt_gpu_blocked_11_expN_5_Sheet["C" + str(2)].value)/5
print('path length rrt_blocked_gpu')
print(avrage_rrt_gpu_blocked_cost)

data_rrt_star_complex_pro_11_expN_1= openpyxl.load_workbook('data_rrt_star_complex_pro_11_expN_1.xlsx')
data_rrt_star_complex_pro_11_expN_2= openpyxl.load_workbook('data_rrt_star_complex_pro_11_expN_2.xlsx')
data_rrt_star_complex_pro_11_expN_3= openpyxl.load_workbook('data_rrt_star_complex_pro_11_expN_3.xlsx')
data_rrt_star_complex_pro_11_expN_4= openpyxl.load_workbook('data_rrt_star_complex_pro_11_expN_4.xlsx')
data_rrt_star_complex_pro_11_expN_5= openpyxl.load_workbook('data_rrt_star_complex_pro_11_expN_5.xlsx')

data_rrt_star_complex_pro_11_expN_1_Sheet=data_rrt_star_complex_pro_11_expN_1.active
data_rrt_star_complex_pro_11_expN_2_Sheet=data_rrt_star_complex_pro_11_expN_2.active
data_rrt_star_complex_pro_11_expN_3_Sheet=data_rrt_star_complex_pro_11_expN_3.active
data_rrt_star_complex_pro_11_expN_4_Sheet=data_rrt_star_complex_pro_11_expN_4.active
data_rrt_star_complex_pro_11_expN_5_Sheet=data_rrt_star_complex_pro_11_expN_5.active

data_rrt_star_complex_pro_11_expN_1_Row=data_rrt_star_complex_pro_11_expN_1_Sheet.max_row
data_rrt_star_complex_pro_11_expN_2_Row=data_rrt_star_complex_pro_11_expN_2_Sheet.max_row
data_rrt_star_complex_pro_11_expN_3_Row=data_rrt_star_complex_pro_11_expN_3_Sheet.max_row
data_rrt_star_complex_pro_11_expN_4_Row=data_rrt_star_complex_pro_11_expN_4_Sheet.max_row
data_rrt_star_complex_pro_11_expN_5_Row=data_rrt_star_complex_pro_11_expN_5_Sheet.max_row

avrage_rrt_star_total_time=[]
avrage_rrt_star_cost=(data_rrt_star_complex_pro_11_expN_1_Sheet["C" + str(2)].value+data_rrt_star_complex_pro_11_expN_2_Sheet["C" + str(2)].value+data_rrt_star_complex_pro_11_expN_3_Sheet["C" + str(2)].value+data_rrt_star_complex_pro_11_expN_4_Sheet["C" + str(2)].value+data_rrt_star_complex_pro_11_expN_5_Sheet["C" + str(2)].value)/5
print('path length rrt_star')
print(avrage_rrt_star_cost)

data_rrt_star_gpu_complex_pro_11_expN_1= openpyxl.load_workbook('data_rrt_star_gpu_complex_pro_11_expN_1.xlsx')
data_rrt_star_gpu_complex_pro_11_expN_2= openpyxl.load_workbook('data_rrt_star_gpu_complex_pro_11_expN_2.xlsx')
data_rrt_star_gpu_complex_pro_11_expN_3= openpyxl.load_workbook('data_rrt_star_gpu_complex_pro_11_expN_3.xlsx')
data_rrt_star_gpu_complex_pro_11_expN_4= openpyxl.load_workbook('data_rrt_star_gpu_complex_pro_11_expN_4.xlsx')
data_rrt_star_gpu_complex_pro_11_expN_5= openpyxl.load_workbook('data_rrt_star_gpu_complex_pro_11_expN_5.xlsx')

data_rrt_star_gpu_complex_pro_11_expN_1_Sheet=data_rrt_star_gpu_complex_pro_11_expN_1.active
data_rrt_star_gpu_complex_pro_11_expN_2_Sheet=data_rrt_star_gpu_complex_pro_11_expN_2.active
data_rrt_star_gpu_complex_pro_11_expN_3_Sheet=data_rrt_star_gpu_complex_pro_11_expN_3.active
data_rrt_star_gpu_complex_pro_11_expN_4_Sheet=data_rrt_star_gpu_complex_pro_11_expN_4.active
data_rrt_star_gpu_complex_pro_11_expN_5_Sheet=data_rrt_star_gpu_complex_pro_11_expN_5.active

data_rrt_star_gpu_complex_pro_11_expN_1_Row=data_rrt_star_gpu_complex_pro_11_expN_1_Sheet.max_row
data_rrt_star_gpu_complex_pro_11_expN_2_Row=data_rrt_star_gpu_complex_pro_11_expN_2_Sheet.max_row
data_rrt_star_gpu_complex_pro_11_expN_3_Row=data_rrt_star_gpu_complex_pro_11_expN_3_Sheet.max_row
data_rrt_star_gpu_complex_pro_11_expN_4_Row=data_rrt_star_gpu_complex_pro_11_expN_4_Sheet.max_row
data_rrt_star_gpu_complex_pro_11_expN_5_Row=data_rrt_star_gpu_complex_pro_11_expN_5_Sheet.max_row

avrage_rrt_star_gpu_total_time=[]
avrage_rrt_star_gpu_cost=(data_rrt_star_gpu_complex_pro_11_expN_1_Sheet["C" + str(2)].value+data_rrt_star_gpu_complex_pro_11_expN_2_Sheet["C" + str(2)].value+data_rrt_star_gpu_complex_pro_11_expN_3_Sheet["C" + str(2)].value+data_rrt_star_gpu_complex_pro_11_expN_4_Sheet["C" + str(2)].value+data_rrt_star_gpu_complex_pro_11_expN_5_Sheet["C" + str(2)].value)/5
print('path length rrt_star_gpu')
print(avrage_rrt_star_gpu_cost)

data_rrt_complex_pro_11_expN_1= openpyxl.load_workbook('data_rrt_complex_pro_11_expN_1.xlsx')
data_rrt_complex_pro_11_expN_2= openpyxl.load_workbook('data_rrt_complex_pro_11_expN_2.xlsx')
data_rrt_complex_pro_11_expN_3= openpyxl.load_workbook('data_rrt_complex_pro_11_expN_3.xlsx')
data_rrt_complex_pro_11_expN_4= openpyxl.load_workbook('data_rrt_complex_pro_11_expN_4.xlsx')
data_rrt_complex_pro_11_expN_5= openpyxl.load_workbook('data_rrt_complex_pro_11_expN_5.xlsx')

data_rrt_complex_pro_11_expN_1_Sheet=data_rrt_complex_pro_11_expN_1.active
data_rrt_complex_pro_11_expN_2_Sheet=data_rrt_complex_pro_11_expN_2.active
data_rrt_complex_pro_11_expN_3_Sheet=data_rrt_complex_pro_11_expN_3.active
data_rrt_complex_pro_11_expN_4_Sheet=data_rrt_complex_pro_11_expN_4.active
data_rrt_complex_pro_11_expN_5_Sheet=data_rrt_complex_pro_11_expN_5.active

data_rrt_complex_pro_11_expN_1_Row=data_rrt_complex_pro_11_expN_1_Sheet.max_row
data_rrt_complex_pro_11_expN_2_Row=data_rrt_complex_pro_11_expN_2_Sheet.max_row
data_rrt_complex_pro_11_expN_3_Row=data_rrt_complex_pro_11_expN_3_Sheet.max_row
data_rrt_complex_pro_11_expN_4_Row=data_rrt_complex_pro_11_expN_4_Sheet.max_row
data_rrt_complex_pro_11_expN_5_Row=data_rrt_complex_pro_11_expN_5_Sheet.max_row


avrage_rrt_complex_pro_cost=(data_rrt_complex_pro_11_expN_1_Sheet["C" + str(2)].value+data_rrt_complex_pro_11_expN_2_Sheet["C" + str(2)].value+data_rrt_complex_pro_11_expN_3_Sheet["C" + str(2)].value+data_rrt_complex_pro_11_expN_4_Sheet["C" + str(2)].value+data_rrt_complex_pro_11_expN_5_Sheet["C" + str(2)].value)/5

avrage_rrt_complex_pro_total=(data_rrt_complex_pro_11_expN_1_Sheet["B" + str(data_rrt_complex_pro_11_expN_1_Row)].value+data_rrt_complex_pro_11_expN_2_Sheet["B" + str(data_rrt_complex_pro_11_expN_2_Row)].value+data_rrt_complex_pro_11_expN_3_Sheet["B" + str(data_rrt_complex_pro_11_expN_3_Row)].value+data_rrt_complex_pro_11_expN_4_Sheet["B" + str(data_rrt_complex_pro_11_expN_4_Row)].value+data_rrt_complex_pro_11_expN_5_Sheet["B" + str(data_rrt_complex_pro_11_expN_5_Row)].value)/5

print('Average time for RRT')
print(avrage_rrt_complex_pro_total)
print('Average path length for RRT')
print(avrage_rrt_complex_pro_cost)

data_rrt_gpu_complex_pro_11_expN_1= openpyxl.load_workbook('data_rrt_gpu_complex_pro_11_expN_1.xlsx')
data_rrt_gpu_complex_pro_11_expN_2= openpyxl.load_workbook('data_rrt_gpu_complex_pro_11_expN_2.xlsx')
data_rrt_gpu_complex_pro_11_expN_3= openpyxl.load_workbook('data_rrt_gpu_complex_pro_11_expN_3.xlsx')
data_rrt_gpu_complex_pro_11_expN_4= openpyxl.load_workbook('data_rrt_gpu_complex_pro_11_expN_4.xlsx')
data_rrt_gpu_complex_pro_11_expN_5= openpyxl.load_workbook('data_rrt_gpu_complex_pro_11_expN_5.xlsx')

data_rrt_gpu_complex_pro_11_expN_1_Sheet=data_rrt_gpu_complex_pro_11_expN_1.active
data_rrt_gpu_complex_pro_11_expN_2_Sheet=data_rrt_gpu_complex_pro_11_expN_2.active
data_rrt_gpu_complex_pro_11_expN_3_Sheet=data_rrt_gpu_complex_pro_11_expN_3.active
data_rrt_gpu_complex_pro_11_expN_4_Sheet=data_rrt_gpu_complex_pro_11_expN_4.active
data_rrt_gpu_complex_pro_11_expN_5_Sheet=data_rrt_gpu_complex_pro_11_expN_5.active

data_rrt_gpu_complex_pro_11_expN_1_Row=data_rrt_gpu_complex_pro_11_expN_1_Sheet.max_row
data_rrt_gpu_complex_pro_11_expN_2_Row=data_rrt_gpu_complex_pro_11_expN_2_Sheet.max_row
data_rrt_gpu_complex_pro_11_expN_3_Row=data_rrt_gpu_complex_pro_11_expN_3_Sheet.max_row
data_rrt_gpu_complex_pro_11_expN_4_Row=data_rrt_gpu_complex_pro_11_expN_4_Sheet.max_row
data_rrt_gpu_complex_pro_11_expN_5_Row=data_rrt_gpu_complex_pro_11_expN_5_Sheet.max_row


avrage_rrt_gpu_complex_pro_cost=(data_rrt_gpu_complex_pro_11_expN_1_Sheet["C" + str(2)].value+data_rrt_gpu_complex_pro_11_expN_2_Sheet["C" + str(2)].value+data_rrt_gpu_complex_pro_11_expN_3_Sheet["C" + str(2)].value+data_rrt_gpu_complex_pro_11_expN_4_Sheet["C" + str(2)].value+data_rrt_gpu_complex_pro_11_expN_5_Sheet["C" + str(2)].value)/5

avrage_rrt_gpu_complex_pro_total=(data_rrt_gpu_complex_pro_11_expN_1_Sheet["B" + str(data_rrt_gpu_complex_pro_11_expN_1_Row)].value+data_rrt_gpu_complex_pro_11_expN_2_Sheet["B" + str(data_rrt_gpu_complex_pro_11_expN_2_Row)].value+data_rrt_gpu_complex_pro_11_expN_3_Sheet["B" + str(data_rrt_gpu_complex_pro_11_expN_3_Row)].value+data_rrt_gpu_complex_pro_11_expN_4_Sheet["B" + str(data_rrt_gpu_complex_pro_11_expN_4_Row)].value+data_rrt_gpu_complex_pro_11_expN_5_Sheet["B" + str(data_rrt_gpu_complex_pro_11_expN_5_Row)].value)/5

print('Average time for RRT on GPU')
print(avrage_rrt_gpu_complex_pro_total)
print('Average path length for RRT on GPU')
print(avrage_rrt_gpu_complex_pro_cost)
std1=[]
std2=[]
std3=[]
std4=[]
for i in range(2, data_rrt_blocked_11_expN_1_Row):
    
    avrage_rrt_blocked_total_time.append((data_rrt_blocked_11_expN_1_Sheet["B" + str(i)].value+data_rrt_blocked_11_expN_2_Sheet["B" + str(i)].value+data_rrt_blocked_11_expN_3_Sheet["B" + str(i)].value+data_rrt_blocked_11_expN_4_Sheet["B" + str(i)].value+data_rrt_blocked_11_expN_5_Sheet["B" + str(i)].value)/5)
    a1=[data_rrt_blocked_11_expN_1_Sheet["B" + str(i)].value,data_rrt_blocked_11_expN_2_Sheet["B" + str(i)].value,data_rrt_blocked_11_expN_3_Sheet["B" + str(i)].value,data_rrt_blocked_11_expN_4_Sheet["B" + str(i)].value,data_rrt_blocked_11_expN_5_Sheet["B" + str(i)].value]
    std1.append(statistics.stdev(a1))
    
    avrage_rrt_gpu_blocked_total_time.append((data_rrt_gpu_blocked_11_expN_1_Sheet["B" + str(i)].value+data_rrt_gpu_blocked_11_expN_2_Sheet["B" + str(i)].value+data_rrt_gpu_blocked_11_expN_3_Sheet["B" + str(i)].value+data_rrt_gpu_blocked_11_expN_4_Sheet["B" + str(i)].value+data_rrt_gpu_blocked_11_expN_5_Sheet["B" + str(i)].value)/5)
    a2=[data_rrt_gpu_blocked_11_expN_1_Sheet["B" + str(i)].value,data_rrt_gpu_blocked_11_expN_2_Sheet["B" + str(i)].value,data_rrt_gpu_blocked_11_expN_3_Sheet["B" + str(i)].value,data_rrt_gpu_blocked_11_expN_4_Sheet["B" + str(i)].value,data_rrt_gpu_blocked_11_expN_5_Sheet["B" + str(i)].value]
    std2.append(statistics.stdev(a2))
    
    avrage_rrt_star_total_time.append((data_rrt_star_complex_pro_11_expN_1_Sheet["B" + str(i)].value+data_rrt_star_complex_pro_11_expN_2_Sheet["B" + str(i)].value+data_rrt_star_complex_pro_11_expN_3_Sheet["B" + str(i)].value+data_rrt_star_complex_pro_11_expN_4_Sheet["B" + str(i)].value+data_rrt_star_complex_pro_11_expN_5_Sheet["B" + str(i)].value)/5)
    a3=[data_rrt_star_complex_pro_11_expN_1_Sheet["B" + str(i)].value,data_rrt_star_complex_pro_11_expN_2_Sheet["B" + str(i)].value,data_rrt_star_complex_pro_11_expN_3_Sheet["B" + str(i)].value,data_rrt_star_complex_pro_11_expN_4_Sheet["B" + str(i)].value,data_rrt_star_complex_pro_11_expN_5_Sheet["B" + str(i)].value]
    std3.append(statistics.stdev(a3))
    
    avrage_rrt_star_gpu_total_time.append((data_rrt_star_gpu_complex_pro_11_expN_1_Sheet["B" + str(i)].value+data_rrt_star_gpu_complex_pro_11_expN_2_Sheet["B" + str(i)].value+data_rrt_star_gpu_complex_pro_11_expN_3_Sheet["B" + str(i)].value+data_rrt_star_gpu_complex_pro_11_expN_4_Sheet["B" + str(i)].value+data_rrt_star_gpu_complex_pro_11_expN_5_Sheet["B" + str(i)].value)/5)
    a4=[data_rrt_star_gpu_complex_pro_11_expN_1_Sheet["B" + str(i)].value,data_rrt_star_gpu_complex_pro_11_expN_2_Sheet["B" + str(i)].value,data_rrt_star_gpu_complex_pro_11_expN_3_Sheet["B" + str(i)].value,data_rrt_star_gpu_complex_pro_11_expN_4_Sheet["B" + str(i)].value,data_rrt_star_gpu_complex_pro_11_expN_5_Sheet["B" + str(i)].value]
    std4.append(statistics.stdev(a4))
    
print('total time RRT blocked')
print(np.array(avrage_rrt_blocked_total_time[len(np.array(avrage_rrt_blocked_total_time))-1]))
print('total time RRT blocked GPU')
print(np.array(avrage_rrt_gpu_blocked_total_time[len(np.array(avrage_rrt_gpu_blocked_total_time))-1]))
print('total time RRT Star')
print(np.array(avrage_rrt_star_total_time[len(np.array(avrage_rrt_star_total_time))-1]))
print('total time RRT Star GPU')
print(np.array(avrage_rrt_star_gpu_total_time[len(np.array(avrage_rrt_star_gpu_total_time))-1]))

x1=np.empty_like(avrage_rrt_blocked_total_time)
x1=range(0,len(avrage_rrt_blocked_total_time))
plt.plot(x1, avrage_rrt_blocked_total_time, label = "RRT Blocked")

#print('std1')
#print(std1)
plt.fill_between(x1, np.array(avrage_rrt_blocked_total_time)-np.array(std1), np.array(avrage_rrt_blocked_total_time)+np.array(std1),alpha=0.3)
# line 2 points
# plotting the line 2 points 
x2=np.empty_like(avrage_rrt_gpu_blocked_total_time)
x2=range(0,len(avrage_rrt_gpu_blocked_total_time))
plt.plot(x2, avrage_rrt_gpu_blocked_total_time, label = "RRT Blocked GPU")


plt.fill_between(x2, np.array(avrage_rrt_gpu_blocked_total_time)-np.array(std2), np.array(avrage_rrt_gpu_blocked_total_time)+np.array(std2),alpha=0.3)

# line 3 points
# plotting the line 3 points 
x3=np.empty_like(avrage_rrt_star_total_time)
x3=range(0,len(avrage_rrt_star_total_time))
plt.plot(x3, avrage_rrt_star_total_time, label = "RRT Star")

plt.fill_between(x3, np.array(avrage_rrt_star_total_time)-np.array(std3), np.array(avrage_rrt_star_total_time)+np.array(std3),alpha=0.5)
# line 4 points
# plotting the line 4 points 
x4=np.empty_like(avrage_rrt_star_gpu_total_time)
x4=range(0,len(avrage_rrt_star_gpu_total_time))
plt.plot(x4, avrage_rrt_star_gpu_total_time, label = "RRT Star GPU")

plt.fill_between(x4, np.array(avrage_rrt_star_gpu_total_time)-np.array(std4), np.array(avrage_rrt_star_gpu_total_time)+np.array(std4),alpha=0.5)
plt.ylabel('Total time(sec)')
plt.xlabel('Iteration')
plt.title('Total Time vs Iteration')
# show a legend on the plot
plt.legend()
  
# function to show the plot
plt.show()
